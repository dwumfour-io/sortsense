"""
SortSense Text Extraction Module

Handles text extraction from various file formats using OCR,
pdftotext, python-docx, and other libraries.
Includes computer vision support for image classification.
"""

import logging
import os
import subprocess
import tempfile
from typing import Optional, Tuple

from sortsense.config import Config

logger = logging.getLogger(__name__)


class TextExtractor:
    """Handles text extraction from various file types"""
    
    def __init__(self, config: Optional[Config] = None, use_vision: bool = False):
        self.config = config or Config()
        self.use_vision = use_vision
        self.vision_analyzer = None
        
        if use_vision:
            self._init_vision()
    
    def _init_vision(self) -> None:
        """Initialize vision analyzer if available"""
        try:
            from sortsense.vision import VisionAnalyzer
            # Pass user's vision categories if available
            vision_cats = self.config.vision_categories if hasattr(self.config, 'vision_categories') else None
            self.vision_analyzer = VisionAnalyzer(categories=vision_cats)
            if self.vision_analyzer.is_available():
                logger.info("Vision analysis enabled")
            else:
                logger.warning("Vision dependencies not available")
                self.vision_analyzer = None
        except Exception as e:
            logger.warning(f"Could not initialize vision: {e}")
            self.vision_analyzer = None
    
    def extract(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from a file based on its extension.
        
        Args:
            filepath: Path to the file
            
        Returns:
            Tuple of (extraction_method, extracted_text)
            Methods: 'ocr', 'text', 'pdf', 'docx', 'xlsx', 'vision', 'filename', 'skip', 'error'
        """
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext in self.config.pdf_extensions:
            return self.extract_from_pdf(filepath)
        elif ext in self.config.image_extensions:
            return self.extract_from_image(filepath)
        elif ext in self.config.word_extensions:
            return self.extract_from_docx(filepath)
        elif ext in self.config.excel_extensions:
            return self.extract_from_xlsx(filepath)
        elif ext in self.config.text_extensions:
            return self.extract_from_text(filepath)
        else:
            return 'filename', ''
    
    def extract_with_vision(self, filepath: str) -> Tuple[str, float, list]:
        """
        Analyze image using computer vision.
        
        Returns:
            Tuple of (category, confidence, matched_descriptions)
        """
        if not self.vision_analyzer:
            return "unsorted", 0.0, []
        
        ext = os.path.splitext(filepath)[1].lower()
        if ext not in self.config.image_extensions:
            return "unsorted", 0.0, []
        
        return self.vision_analyzer.analyze_image(filepath)
    def extract_from_pdf(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from PDF using pdftotext, fallback to OCR.
        
        Args:
            filepath: Path to PDF file
            
        Returns:
            Tuple of (method, text)
        """
        if not self.config.pdftotext_path:
            logger.warning("pdftotext not available, trying OCR")
            return self._ocr_pdf(filepath)
        
        try:
            result = subprocess.run(
                [self.config.pdftotext_path, '-l', '2', filepath, '-'],
                capture_output=True,
                text=True,
                timeout=self.config.settings.ocr_timeout
            )
            
            text = result.stdout.strip()
            if text:
                return 'pdf', text[:self.config.settings.max_text_length]
            
            # No text found, try OCR
            return self._ocr_pdf(filepath)
            
        except subprocess.TimeoutExpired:
            logger.error(f"Timeout extracting PDF: {filepath}")
            return 'error', 'Timeout during PDF extraction'
        except Exception as e:
            logger.error(f"Error extracting PDF {filepath}: {e}")
            return 'error', str(e)[:100]
    
    def _ocr_pdf(self, filepath: str) -> Tuple[str, str]:
        """OCR a scanned PDF by converting to image first"""
        if not self.config.pdftoppm_path or not self.config.tesseract_path:
            logger.warning("OCR tools not available for PDF")
            return 'skip', ''
        
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                img_prefix = os.path.join(tmpdir, 'page')
                
                # Convert first page to PNG
                subprocess.run(
                    [self.config.pdftoppm_path, '-png', '-f', '1', '-l', '1', filepath, img_prefix],
                    capture_output=True,
                    timeout=self.config.settings.ocr_timeout
                )
                
                # Find the generated image
                png_file = f"{img_prefix}-1.png"
                if not os.path.exists(png_file):
                    png_file = f"{img_prefix}-01.png"
                
                if os.path.exists(png_file):
                    return self.extract_from_image(png_file)
                
                return 'empty', ''
                
        except subprocess.TimeoutExpired:
            logger.error(f"Timeout OCR-ing PDF: {filepath}")
            return 'error', 'Timeout during OCR'
        except Exception as e:
            logger.error(f"Error OCR-ing PDF {filepath}: {e}")
            return 'error', str(e)[:100]
    
    def extract_from_image(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from image using Tesseract OCR.
        
        Args:
            filepath: Path to image file
            
        Returns:
            Tuple of (method, text)
        """
        if not self.config.tesseract_path:
            logger.warning("Tesseract not available for image OCR")
            return 'skip', ''
        
        try:
            result = subprocess.run(
                [self.config.tesseract_path, filepath, 'stdout'],
                capture_output=True,
                text=True,
                timeout=self.config.settings.ocr_timeout
            )
            
            text = result.stdout.strip()
            return 'ocr', text[:self.config.settings.max_text_length]
            
        except subprocess.TimeoutExpired:
            logger.error(f"Timeout OCR-ing image: {filepath}")
            return 'error', 'OCR timeout'
        except FileNotFoundError:
            logger.error("Tesseract not found")
            return 'error', 'Tesseract not found'
        except Exception as e:
            logger.error(f"Error OCR-ing image {filepath}: {e}")
            return 'error', str(e)[:100]
    
    def extract_from_docx(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from Word document using python-docx.
        
        Args:
            filepath: Path to .docx file
            
        Returns:
            Tuple of (method, text)
        """
        try:
            from docx import Document
            
            doc = Document(filepath)
            text_parts = []
            
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)
            
            text = '\n'.join(text_parts)
            return 'docx', text[:self.config.settings.max_text_length]
            
        except ImportError:
            logger.warning("python-docx not installed, skipping .docx extraction")
            return 'skip', ''
        except Exception as e:
            logger.error(f"Error extracting DOCX {filepath}: {e}")
            return 'error', str(e)[:100]
    
    def extract_from_xlsx(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from Excel file using openpyxl.
        
        Args:
            filepath: Path to .xlsx file
            
        Returns:
            Tuple of (method, text)
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath, read_only=True, data_only=True)
            text_parts = []
            
            for sheet in wb.worksheets[:2]:  # First 2 sheets only
                for row in sheet.iter_rows(max_row=50):  # First 50 rows
                    for cell in row:
                        if cell.value:
                            text_parts.append(str(cell.value))
            
            wb.close()
            text = ' '.join(text_parts)
            return 'xlsx', text[:self.config.settings.max_text_length]
            
        except ImportError:
            logger.warning("openpyxl not installed, skipping .xlsx extraction")
            return 'skip', ''
        except Exception as e:
            logger.error(f"Error extracting XLSX {filepath}: {e}")
            return 'error', str(e)[:100]
    
    def extract_from_text(self, filepath: str) -> Tuple[str, str]:
        """
        Extract text from plain text files.
        
        Args:
            filepath: Path to text file
            
        Returns:
            Tuple of (method, text)
        """
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                text = f.read()
            return 'text', text[:self.config.settings.max_text_length]
        except Exception as e:
            logger.error(f"Error reading text file {filepath}: {e}")
            return 'error', str(e)[:100]
